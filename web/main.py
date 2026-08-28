import io
import json
import os
import shutil
from datetime import datetime
from typing import Dict, List
from urllib.parse import unquote_plus

from fastapi import (
    Depends,
    FastAPI,
    Form,
    HTTPException,
    Query,
    Request,
    status,
)
from fastapi.responses import HTMLResponse, PlainTextResponse, RedirectResponse, StreamingResponse
from fastapi.security import HTTPBasic, HTTPBasicCredentials
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates

from jinja2 import Environment, FileSystemLoader
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from cocktail_calc.calculator import (
    _find_pdf_font,
    calculate_shopping_list,
    format_report,
    generate_pdf_report,
    generate_txt_report,
    parse_order,
)
from cocktail_calc.config import DATA_DIR, DB_FILE
from cocktail_calc import events_repository
from cocktail_calc.repository import load_database, load_default_prices

app = FastAPI(title="CocktailCalc Pro")

app.mount("/static", StaticFiles(directory="web/static"), name="static")
templates = Jinja2Templates(
    env=Environment(
        loader=FileSystemLoader("web/templates"),
        autoescape=True,
        cache_size=0,
    )
)


db = load_database()
prices = load_default_prices()


def reload_db():
    global db, prices
    db = load_database()
    prices = load_default_prices()


# ─── Публичные страницы ──────────────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    return templates.TemplateResponse(
        request=request,
        name="index.html",
        context={
            "cocktails": db.cocktails,
            "cocktail_categories": db.cocktail_categories,
        },
    )


@app.post("/calculate", response_class=HTMLResponse)
async def calculate(request: Request, order: str = Form(...)):
    order_text = order.strip()
    parsed = parse_order(order_text)
    if not parsed:
        raise HTTPException(status_code=400, detail="Не удалось распознать заказ")

    result = calculate_shopping_list(parsed, prices, db)
    report_text = format_report(result, db)

    return templates.TemplateResponse(
        request=request,
        name="report.html",
        context={
            "report_text": report_text,
            "order_text": order_text,
        },
    )


@app.get("/menu", response_class=HTMLResponse)
async def menu(request: Request):
    return templates.TemplateResponse(
        request=request,
        name="menu.html",
        context={
            "cocktails": db.cocktails,
            "cocktail_categories": db.cocktail_categories,
        },
    )


@app.get("/api/calculate")
async def api_calculate(order: str = Query(..., description="Заказ вида 'Кокосовый ром 10, Мохито 5'")):
    parsed = parse_order(unquote_plus(order))
    if not parsed:
        raise HTTPException(status_code=400, detail="Не удалось распознать заказ")
    result = calculate_shopping_list(parsed, prices, db)
    return result


@app.get("/export/txt", response_class=PlainTextResponse)
async def export_txt(order: str = Query(...)):
    parsed = parse_order(unquote_plus(order))
    if not parsed:
        raise HTTPException(status_code=400, detail="Не удалось распознать заказ")
    result = calculate_shopping_list(parsed, prices, db)
    report = generate_txt_report(result, db)
    filename = f"cocktail_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
    return PlainTextResponse(
        content=report,
        media_type="text/plain; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.get("/export/pdf")
async def export_pdf(order: str = Query(...)):
    parsed = parse_order(unquote_plus(order))
    if not parsed:
        raise HTTPException(status_code=400, detail="Не удалось распознать заказ")
    result = calculate_shopping_list(parsed, prices, db)
    pdf_bytes = generate_pdf_report(result, db)
    filename = f"cocktail_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ─── Админка для базы данных ──────────────────────────────────────────────────

security = HTTPBasic()

ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "admin"


def verify_admin(credentials: HTTPBasicCredentials = Depends(security)):
    is_valid = (
        credentials.username == ADMIN_USERNAME
        and credentials.password == ADMIN_PASSWORD
    )
    if not is_valid:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Неверный логин или пароль",
            headers={"WWW-Authenticate": "Basic"},
        )
    return credentials


@app.get("/admin", response_class=HTMLResponse)
async def admin_page(request: Request, credentials=Depends(verify_admin)):
    return templates.TemplateResponse(
        request=request,
        name="admin.html",
        context={"db_file": DB_FILE},
    )


@app.get("/admin/api/db")
async def admin_get_db(credentials=Depends(verify_admin)):
    with open(DB_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


@app.post("/admin/api/db")
async def admin_save_db(data: dict, credentials=Depends(verify_admin)):
    global db, prices

    # Автобэкап перед сохранением
    backup_dir = os.path.join(DATA_DIR, "backups")
    os.makedirs(backup_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = os.path.join(backup_dir, f"cocktails_db_backup_{timestamp}.json")
    shutil.copy2(DB_FILE, backup_path)
    _rotate_backups()

    # Сохраняем новую базу
    with open(DB_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    # Перезагружаем глобальные объекты
    reload_db()

    return {"status": "ok", "backup": backup_path}


@app.get("/admin/api/ingredients")
async def admin_get_ingredients(credentials=Depends(verify_admin)):
    keys = set(db.categories.keys()) | set(db.prices.keys()) | set(db.bottle_volumes.keys())
    result = []
    for key in sorted(keys):
        info = db.ingredient_info.get(key, {})
        category = db.categories.get(key, "")
        result.append({
            "key": key,
            "display_name": info.get("display_name", key),
            "unit": info.get("unit", _default_unit(category)),
            "category": category,
            "volume": db.bottle_volumes.get(key, 0),
            "price": db.prices.get(key, 0),
        })
    return result


@app.get("/admin/api/semi-products")
async def admin_get_semi_products(credentials=Depends(verify_admin)):
    return [
        {
            "key": key,
            "name": pf.name,
            "unit": pf.unit,
            "output_volume": pf.output_volume,
        }
        for key, pf in sorted(db.semi_products.items())
    ]


def _rotate_backups(keep: int = 20):
    backup_dir = os.path.join(DATA_DIR, "backups")
    if not os.path.isdir(backup_dir):
        return
    files = sorted(
        [f for f in os.listdir(backup_dir) if f.startswith("cocktails_db_backup_") and f.endswith(".json")],
        key=lambda f: os.path.getmtime(os.path.join(backup_dir, f)),
        reverse=True,
    )
    for old_file in files[keep:]:
        os.remove(os.path.join(backup_dir, old_file))


def _default_unit(category: str) -> str:
    mapping = {
        "алкоголь": "л",
        "безалкогольное": "л",
        "сироп": "л",
        "пюре": "л",
        "концентрат": "л",
        "сухой_гр": "кг",
        "лёд_кубик": "кг",
        "лёд_фигурный": "шт",
        "украшение_шт": "шт",
        "украшение_гр": "гр",
        "посуда": "шт",
    }
    return mapping.get(category, "л")


def _rename_in_recipes(data: dict, old_key: str, new_key: str):
    """Заменяет ключ ингредиента во всех рецептах коктейлей и ПФ."""
    for section in ("semi_products", "cocktails"):
        for item in data.get(section, {}).values():
            recipe = item.get("recipe", {})
            if old_key in recipe:
                recipe[new_key] = recipe.pop(old_key)
            for sub_key in ("decorations", "glassware"):
                sub = item.get(sub_key, {})
                if old_key in sub:
                    sub[new_key] = sub.pop(old_key)


@app.get("/admin/api/glassware")
async def admin_get_glassware(credentials=Depends(verify_admin)):
    result = []
    for key, category in db.categories.items():
        if category != "посуда":
            continue
        info = db.ingredient_info.get(key, {})
        result.append({
            "key": key,
            "display_name": info.get("display_name", key),
        })
    return sorted(result, key=lambda x: x["key"])


@app.post("/admin/api/ingredient/save")
async def admin_save_ingredient(
    payload: dict,
    credentials=Depends(verify_admin),
):
    global db, prices
    old_key = payload.get("old_key", "").strip()
    new_key = payload.get("new_key", "").strip()
    if not old_key or not new_key:
        raise HTTPException(status_code=400, detail="old_key и new_key обязательны")

    with open(DB_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)

    if old_key not in data.get("categories", {}):
        raise HTTPException(status_code=400, detail="Ингредиент не найден")

    if new_key != old_key and new_key in data.get("categories", {}):
        raise HTTPException(status_code=400, detail="Новый ключ уже занят")

    # Автобэкап перед изменением
    backup_dir = os.path.join(DATA_DIR, "backups")
    os.makedirs(backup_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = os.path.join(backup_dir, f"cocktails_db_backup_{timestamp}.json")
    shutil.copy2(DB_FILE, backup_path)
    _rotate_backups()

    # Переименовываем ключи в основных секциях
    for section in ("categories", "prices", "bottle_volumes"):
        if old_key in data.get(section, {}):
            data[section][new_key] = data[section].pop(old_key)

    if "ingredient_info" not in data:
        data["ingredient_info"] = {}

    if old_key in data["ingredient_info"]:
        data["ingredient_info"][new_key] = data["ingredient_info"].pop(old_key)
    else:
        data["ingredient_info"][new_key] = {}

    # Обновляем поля
    info = data["ingredient_info"][new_key]
    info["display_name"] = payload.get("display_name", new_key).strip() or new_key
    info["unit"] = payload.get("unit", _default_unit(data["categories"].get(new_key, ""))).strip()

    data["categories"][new_key] = payload.get("category", data["categories"].get(new_key, "")).strip()
    data["prices"][new_key] = int(payload.get("price", data["prices"].get(new_key, 0)))
    data["bottle_volumes"][new_key] = float(payload.get("volume", data["bottle_volumes"].get(new_key, 0)))

    # Переименовываем во всех рецептах
    if new_key != old_key:
        _rename_in_recipes(data, old_key, new_key)

    with open(DB_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    reload_db()
    prices = load_default_prices()

    return {"status": "ok", "backup": backup_path}


@app.post("/admin/api/backup")
async def admin_backup(credentials=Depends(verify_admin)):
    backup_dir = os.path.join(DATA_DIR, "backups")
    os.makedirs(backup_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = os.path.join(backup_dir, f"cocktails_db_backup_{timestamp}.json")
    shutil.copy2(DB_FILE, backup_path)
    _rotate_backups()
    return {"status": "ok", "backup": backup_path}


def _format_amount_and_unit(name: str, amount: float, category: str = "") -> tuple:
    unit = _default_unit(category) if category else "л"
    if "лёд" in name or category in ("лёд_кубик", "лёд_фигурный"):
        unit = "шт" if "фигурный" in name or category == "лёд_фигурный" else "кг"
        return round(amount, 3), unit
    if category == "сухой_гр":
        return round(amount, 3), "кг"
    if category == "украшение_гр":
        return round(amount, 3), "гр"
    if category == "украшение_шт":
        return int(amount), "шт"
    if category == "посуда":
        return int(amount), "шт"
    if isinstance(amount, float) and 0 < amount < 0.1:
        return round(amount * 1000, 2), "мл"
    return round(amount, 3), "л"


def _generate_ttk_excel(portions: int = 1):
    wb = Workbook()
    ws = wb.active
    ws.title = "ТТК Коктейли"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4BACC6")
    cat_fill = PatternFill("solid", fgColor="9BBB59")
    label_fill = PatternFill("solid", fgColor="F79646")
    item_odd_fill = PatternFill("solid", fgColor="D9D9D9")
    item_even_fill = PatternFill("solid", fgColor="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def append_row(col1, name, per_one, portions_val, total, fill=None, bold=False, font_color="000000"):
        ws.append([col1, name, per_one, portions_val, total])
        row = ws.max_row
        for c in range(1, 6):
            cell = ws.cell(row=row, column=c)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if fill:
                cell.fill = fill
            if bold:
                cell.font = Font(bold=True, color=font_color)
            if c == 1 and col1:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Шапка
    append_row("", "Наименование", "Количество на 1 коктейль", "Количество порций", "Общее количество (Кг/Л)", fill=header_fill, bold=True, font_color="FFFFFF")

    for idx, (ckey, cocktail) in enumerate(db.cocktails.items(), 1):
        name = cocktail.name
        recipe = cocktail.recipe
        decorations = cocktail.decorations
        glassware = cocktail.glassware

        # Выход коктейля — сумма жидких ингредиентов
        liquid_total = 0.0
        for ing, amount in recipe.items():
            if "лёд" in ing:
                continue
            cat = db.categories.get(ing, "")
            if cat in ("алкоголь", "безалкогольное", "сироп", "пюре", "концентрат", ""):
                liquid_total += amount
        cocktail_per_one = round(liquid_total, 3)
        cocktail_total = round(cocktail_per_one * portions, 3)

        # Заголовок коктейля
        fill = cat_fill if idx % 2 == 1 else PatternFill("solid", fgColor="B7D58C")
        append_row("Коктейль:", name, cocktail_per_one, portions, cocktail_total, fill=fill, bold=True)

        # Ингредиенты рецепта
        recipe_items = [(ing, amount) for ing, amount in recipe.items() if "лёд" not in ing]
        for i, (ing, amount) in enumerate(recipe_items):
            cat = db.categories.get(ing, "")
            a, u = _format_amount_and_unit(ing, amount, cat)
            t = round(a * portions, 3) if u != "шт" else int(a) * portions
            item_fill = item_odd_fill if i % 2 == 0 else item_even_fill
            append_row("", ing, a, portions, t, fill=item_fill)

        # Лёд
        ice_items = [(ing, amount) for ing, amount in recipe.items() if "лёд" in ing]
        if ice_items:
            for i, (ing, amount) in enumerate(ice_items):
                cat = db.categories.get(ing, "")
                a, u = _format_amount_and_unit(ing, amount, cat)
                t = round(a * portions, 3) if u != "шт" else int(a) * portions
                label = "Лёд:" if i == 0 else ""
                append_row(label, ing, a, portions, t, fill=label_fill, bold=True)
        else:
            append_row("Лёд:", "Отсутствует", "", portions, "", fill=label_fill, bold=True)

        # Украшение
        if decorations:
            for i, (d, amount) in enumerate(decorations.items()):
                cat = db.categories.get(d, "")
                a, u = _format_amount_and_unit(d, amount, cat)
                t = round(a * portions, 3) if u != "шт" else int(a) * portions
                label = "Украшение:" if i == 0 else ""
                append_row(label, d, a, portions, t, fill=label_fill, bold=True)
        else:
            append_row("Украшение:", "Отсутствует", "", portions, "", fill=label_fill, bold=True)

        # Посуда
        if glassware:
            for i, (g, amount) in enumerate(glassware.items()):
                a, u = _format_amount_and_unit(g, int(amount), "посуда")
                t = a * portions
                label = "Посуда:" if i == 0 else ""
                append_row(label, g, a, portions, t, fill=label_fill, bold=True)
        else:
            append_row("Посуда:", "Отсутствует", "", portions, "", fill=label_fill, bold=True)

        # Разделитель
        ws.append(["", "", "", "", ""])

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 26
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@app.get("/admin/api/download_ttk")
async def admin_download_ttk(portions: int = 1, credentials=Depends(verify_admin)):
    if portions < 1:
        raise HTTPException(status_code=400, detail="Количество порций должно быть ≥ 1")
    output = _generate_ttk_excel(portions)
    filename = f"ttk_cocktails_{portions}port_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ─── CRM мероприятий ─────────────────────────────────────────────────────────

_BAR_LABELS = {
    "white_columns": "Белый с колоннами",
    "white_no_columns": "Белый без колонн",
    "black": "Чёрный",
    "none": "Нет",
}

_SHELF_LABELS = {
    "black_white": "Чёрный с белыми полками",
    "gold_black": "Золотой с чёрными полками",
    "none": "Нет",
}

_PYRAMID_LABELS = {
    "56": "56",
    "84": "84",
    "120": "120",
    "none": "Нет",
}

_MENU_LABELS = {
    "us": "С нас",
    "client": "С заказчика",
}

_DECORATIONS_LIST = [
    "базовые",
    "клеймо",
    "лёд с интеграцией",
    "печать на съедобной бумаге",
    "трафарет",
    "азот",
    "бластер",
    "кондитерское украшение",
    "шоко трансфер",
    "принтер",
]


def _build_event_data(
    title: str,
    client: str,
    bartenders_count: str,
    date: str,
    address: str,
    departure: str,
    setup: str,
    start: str,
    end: str,
    manager_contact: str,
    bar: str,
    bar_comment: str,
    shelf: str,
    shelf_comment: str,
    pyramid: str,
    pyramid_comment: str,
    decorations: List[str],
    decorations_comment: str,
    clothing: str,
    menu: str,
    comment: str,
    cocktails_json: str,
) -> Dict:
    try:
        bartenders = int(bartenders_count) if bartenders_count.strip() else 0
    except ValueError:
        bartenders = 0

    try:
        cocktails = json.loads(cocktails_json or "{}")
        if not isinstance(cocktails, dict):
            cocktails = {}
    except json.JSONDecodeError:
        cocktails = {}

    return {
        "title": title.strip(),
        "client": client.strip(),
        "bartenders_count": bartenders,
        "date": date,
        "address": address.strip(),
        "timings": {
            "departure": departure,
            "setup": setup,
            "start": start,
            "end": end,
        },
        "manager_contact": manager_contact.strip(),
        "bar": bar,
        "bar_comment": bar_comment.strip(),
        "shelf": shelf,
        "shelf_comment": shelf_comment.strip(),
        "pyramid": pyramid,
        "pyramid_comment": pyramid_comment.strip(),
        "decorations": decorations,
        "decorations_comment": decorations_comment.strip(),
        "clothing": clothing.strip(),
        "menu": menu,
        "comment": comment.strip(),
        "cocktails": cocktails,
    }


def _generate_event_pdf(event: Dict) -> bytes:
    from fpdf import FPDF

    font_path = _find_pdf_font()
    if not font_path:
        raise RuntimeError("Не найден TTF-шрифт с поддержкой кириллицы для PDF")

    font_dir = os.path.dirname(font_path)
    bold_candidates = [
        os.path.join(font_dir, "DejaVuSans-Bold.ttf"),
        os.path.join(font_dir, "LiberationSans-Bold.ttf"),
        os.path.join(font_dir, "arialbd.ttf"),
        os.path.join(font_dir, "calibrib.ttf"),
    ]
    bold_path = font_path
    for cand in bold_candidates:
        if os.path.exists(cand):
            bold_path = cand
            break

    class EventPDF(FPDF):
        def header(self):
            self.set_font("EventFont", "B", 14)
            self.set_text_color(96, 18, 3)
            self.cell(0, 10, event.get("title", "Мероприятие"), ln=True, align="C")
            self.set_font("EventFont", "", 9)
            self.set_text_color(110, 110, 115)
            self.cell(0, 6, f"Дата: {event.get('date') or '—'}", ln=True, align="C")
            self.ln(4)

        def footer(self):
            self.set_y(-15)
            self.set_font("EventFont", "", 8)
            self.set_text_color(110, 110, 115)
            self.cell(0, 10, f"Стр. {self.page_no()}", align="C")

    pdf = EventPDF()
    pdf.add_font("EventFont", "", font_path, uni=True)
    pdf.add_font("EventFont", "B", bold_path, uni=True)
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    def _row(label, value):
        pdf.set_font("EventFont", "B", 10)
        pdf.set_text_color(96, 18, 3)
        pdf.cell(70, 7, label, ln=0)
        pdf.set_font("EventFont", "", 10)
        pdf.set_text_color(29, 29, 31)
        pdf.multi_cell(0, 7, value or "—")

    pdf.set_font("EventFont", "B", 11)
    pdf.set_text_color(96, 18, 3)
    pdf.cell(0, 8, "Основное", ln=True)
    _row("Заказчик:", event.get("client", ""))
    _row("Барменов:", str(event.get("bartenders_count", "")))
    _row("Адрес:", event.get("address", ""))
    _row("Контакт менеджера:", event.get("manager_contact", ""))
    pdf.ln(4)

    pdf.set_font("EventFont", "B", 11)
    pdf.set_text_color(96, 18, 3)
    pdf.cell(0, 8, "Тайминг", ln=True)
    timings = event.get("timings", {})
    _row("Выезд:", timings.get("departure", ""))
    _row("Монтаж:", timings.get("setup", ""))
    _row("Начало:", timings.get("start", ""))
    _row("Финал:", timings.get("end", ""))
    pdf.ln(4)

    pdf.set_font("EventFont", "B", 11)
    pdf.set_text_color(96, 18, 3)
    pdf.cell(0, 8, "Оборудование", ln=True)
    bar = event.get("bar", "none")
    _row("Бар:", f"{_BAR_LABELS.get(bar, bar)}{', ' + event.get('bar_comment', '') if event.get('bar_comment') else ''}")
    shelf = event.get("shelf", "none")
    _row("Стеллаж:", f"{_SHELF_LABELS.get(shelf, shelf)}{', ' + event.get('shelf_comment', '') if event.get('shelf_comment') else ''}")
    pyramid = event.get("pyramid", "none")
    _row("Пирамида:", f"{_PYRAMID_LABELS.get(pyramid, pyramid)}{', ' + event.get('pyramid_comment', '') if event.get('pyramid_comment') else ''}")
    pdf.ln(4)

    pdf.set_font("EventFont", "B", 11)
    pdf.set_text_color(96, 18, 3)
    pdf.cell(0, 8, "Украшения", ln=True)
    decorations = event.get("decorations", [])
    if decorations:
        pdf.set_font("EventFont", "", 10)
        pdf.set_text_color(29, 29, 31)
        for dec in decorations:
            pdf.cell(0, 6, f"• {dec[:1].upper()}{dec[1:]}", ln=True)
    else:
        pdf.set_font("EventFont", "", 10)
        pdf.cell(0, 6, "—", ln=True)
    if event.get("decorations_comment"):
        pdf.set_font("EventFont", "", 9)
        pdf.set_text_color(110, 110, 115)
        pdf.multi_cell(0, 6, event.get("decorations_comment", ""))
    pdf.ln(4)

    pdf.set_font("EventFont", "B", 11)
    pdf.set_text_color(96, 18, 3)
    pdf.cell(0, 8, "Дополнительно", ln=True)
    _row("Форма одежды:", event.get("clothing", ""))
    menu = event.get("menu", "us")
    _row("Меню:", _MENU_LABELS.get(menu, menu))
    _row("Комментарий:", event.get("comment", ""))
    pdf.ln(4)

    cocktails = event.get("cocktails", {})
    if cocktails:
        pdf.set_font("EventFont", "B", 11)
        pdf.set_text_color(96, 18, 3)
        pdf.cell(0, 8, "Коктейли", ln=True)
        total = 0
        for name, qty in sorted(cocktails.items()):
            pdf.set_font("EventFont", "", 10)
            pdf.set_text_color(29, 29, 31)
            pdf.cell(90, 6, name, ln=0)
            pdf.cell(0, 6, f"{qty} шт.", ln=True)
            total += int(qty)
        pdf.set_font("EventFont", "B", 10)
        pdf.set_text_color(96, 18, 3)
        pdf.cell(90, 7, "Итого:", ln=0)
        pdf.cell(0, 7, f"{total} шт.", ln=True)

    pdf_bytes = pdf.output(dest="S")
    return bytes(pdf_bytes) if isinstance(pdf_bytes, bytearray) else pdf_bytes


@app.get("/events", response_class=HTMLResponse)
async def events_list(request: Request):
    return templates.TemplateResponse(
        request=request,
        name="events/list.html",
        context={"events": events_repository.list_events()},
    )


@app.get("/events/new", response_class=HTMLResponse)
async def events_new_form(request: Request):
    return templates.TemplateResponse(
        request=request,
        name="events/form.html",
        context={
            "event": None,
            "cocktails_json": "{}",
            "cocktails": db.cocktails,
            "cocktail_categories": db.cocktail_categories,
            "decorations_list": _DECORATIONS_LIST,
        },
    )


@app.post("/events/new", response_class=RedirectResponse)
async def events_create(
    title: str = Form(...),
    client: str = Form(default=""),
    bartenders_count: str = Form(default="1"),
    date: str = Form(default=""),
    address: str = Form(default=""),
    departure: str = Form(default=""),
    setup: str = Form(default=""),
    start: str = Form(default=""),
    end: str = Form(default=""),
    manager_contact: str = Form(default=""),
    bar: str = Form(default="none"),
    bar_comment: str = Form(default=""),
    shelf: str = Form(default="none"),
    shelf_comment: str = Form(default=""),
    pyramid: str = Form(default="none"),
    pyramid_comment: str = Form(default=""),
    decorations: List[str] = Form(default=[]),
    decorations_comment: str = Form(default=""),
    clothing: str = Form(default=""),
    menu: str = Form(default="us"),
    comment: str = Form(default=""),
    cocktails_json: str = Form(default="{}"),
):
    data = _build_event_data(
        title, client, bartenders_count, date, address,
        departure, setup, start, end, manager_contact,
        bar, bar_comment, shelf, shelf_comment, pyramid, pyramid_comment,
        decorations, decorations_comment, clothing, menu, comment, cocktails_json,
    )
    event_id = events_repository.create_event(data)
    return RedirectResponse(url=f"/events/{event_id}", status_code=status.HTTP_303_SEE_OTHER)


@app.get("/events/{event_id}", response_class=HTMLResponse)
async def events_detail(request: Request, event_id: str):
    event = events_repository.get_event(event_id)
    if not event:
        raise HTTPException(status_code=404, detail="Мероприятие не найдено")
    total_cocktails = sum(int(q) for q in event.get("cocktails", {}).values())
    return templates.TemplateResponse(
        request=request,
        name="events/detail.html",
        context={
            "event": event,
            "bar_label": _BAR_LABELS.get(event.get("bar", "none"), event.get("bar", "—")),
            "shelf_label": _SHELF_LABELS.get(event.get("shelf", "none"), event.get("shelf", "—")),
            "pyramid_label": _PYRAMID_LABELS.get(event.get("pyramid", "none"), event.get("pyramid", "—")),
            "menu_label": _MENU_LABELS.get(event.get("menu", "us"), event.get("menu", "—")),
            "total_cocktails": total_cocktails,
        },
    )


@app.get("/events/{event_id}/edit", response_class=HTMLResponse)
async def events_edit_form(request: Request, event_id: str):
    event = events_repository.get_event(event_id)
    if not event:
        raise HTTPException(status_code=404, detail="Мероприятие не найдено")
    return templates.TemplateResponse(
        request=request,
        name="events/form.html",
        context={
            "event": event,
            "cocktails_json": json.dumps(event.get("cocktails", {}), ensure_ascii=False),
            "cocktails": db.cocktails,
            "cocktail_categories": db.cocktail_categories,
            "decorations_list": _DECORATIONS_LIST,
        },
    )


@app.post("/events/{event_id}/edit", response_class=RedirectResponse)
async def events_update(
    event_id: str,
    title: str = Form(...),
    client: str = Form(default=""),
    bartenders_count: str = Form(default="1"),
    date: str = Form(default=""),
    address: str = Form(default=""),
    departure: str = Form(default=""),
    setup: str = Form(default=""),
    start: str = Form(default=""),
    end: str = Form(default=""),
    manager_contact: str = Form(default=""),
    bar: str = Form(default="none"),
    bar_comment: str = Form(default=""),
    shelf: str = Form(default="none"),
    shelf_comment: str = Form(default=""),
    pyramid: str = Form(default="none"),
    pyramid_comment: str = Form(default=""),
    decorations: List[str] = Form(default=[]),
    decorations_comment: str = Form(default=""),
    clothing: str = Form(default=""),
    menu: str = Form(default="us"),
    comment: str = Form(default=""),
    cocktails_json: str = Form(default="{}"),
):
    data = _build_event_data(
        title, client, bartenders_count, date, address,
        departure, setup, start, end, manager_contact,
        bar, bar_comment, shelf, shelf_comment, pyramid, pyramid_comment,
        decorations, decorations_comment, clothing, menu, comment, cocktails_json,
    )
    ok = events_repository.update_event(event_id, data)
    if not ok:
        raise HTTPException(status_code=404, detail="Мероприятие не найдено")
    return RedirectResponse(url=f"/events/{event_id}", status_code=status.HTTP_303_SEE_OTHER)


@app.post("/events/{event_id}/delete", response_class=RedirectResponse)
async def events_delete(event_id: str):
    ok = events_repository.delete_event(event_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Мероприятие не найдено")
    return RedirectResponse(url="/events", status_code=status.HTTP_303_SEE_OTHER)


@app.get("/events/{event_id}/pdf")
async def events_pdf(event_id: str):
    event = events_repository.get_event(event_id)
    if not event:
        raise HTTPException(status_code=404, detail="Мероприятие не найдено")
    pdf_bytes = _generate_event_pdf(event)
    filename = f"event_{event.get('title', event_id)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
